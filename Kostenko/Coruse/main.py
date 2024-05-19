from PyQt6.QtCore import *
from PyQt6.QtWidgets import *
import openpyxl
import sys
import time

class TableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return float(self.text()) < float(other.text())
        except ValueError:
            return super().__lt__(other)

class Window(QMainWindow):
    def __init__(self):
        super().__init__()

        self.menu_bar = QMenuBar()

        self.menu = QMenu("Файл")
        action1 = self.menu.addAction("Сохранить")
        action1.triggered.connect(self.export)

        self.menu_bar.addMenu(self.menu)

        self.setMenuBar(self.menu_bar)

    def closeEvent(self, event):
        print("closing")
        widget.DataWindow.close()

        event.accept()

    def export(self):
        filename, filter = QFileDialog.getSaveFileName(self, 'Save file', '','Excel files (*.xlsx)')
        wb = openpyxl.Workbook()

        temp = wb["Sheet"]
        wb.remove(temp)

        print(filename)

        centralWidget = self.centralWidget()

        for sheet, table_widget in centralWidget.temp_sheets.items():
            workingSheet = wb.create_sheet(sheet.title)

            labels = []
            for c in range(table_widget.columnCount()):
                it = table_widget.horizontalHeaderItem(c)
                labels.append(str(c+1) if it is None else it.text())
            print(labels)

            for i in range(len(labels)):
                value = labels[i]
                workingSheet.cell(1, i + 1, value)

            if filename:
                for column in range(table_widget.columnCount()):
                    for row in range(table_widget.rowCount()):
                        try:
                            text = str(table_widget.item(row, column).text())
                            workingSheet.cell(row + 2, column + 1, text)
                        except AttributeError:
                            pass
                        
        wb.save(filename)


class DataWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Add Data")
        self.setFixedSize(500, 500)

        self.Layout = QVBoxLayout()
        self.setLayout(self.Layout)

        self.addButton = QPushButton("Добавить")
        self.addButton.clicked.connect(self.acceptInfo)
        self.Layout.addWidget(self.addButton)
        
        self.inputs :list[QLineEdit] = []
        self.labels :list[QLabel] = []

    def acceptInfo(self):
        widget.currentWidget.insertRow(0)

        for i in range(len(self.labels)):
            print(0, i)
            item = TableWidgetItem(str(self.inputs[i].text()))

            widget.currentWidget.setItem(0 , i, item)

    def load(self):
        print("loading stuff")

        self.Layout.removeWidget(self.addButton)

        labels = []

        for c in range(widget.currentWidget.columnCount()):
            label = widget.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
        
        for l in range(len(labels)):
            label = QLabel(labels[l])
            lineEdit = QLineEdit()
            self.Layout.addWidget(label)
            self.Layout.addWidget(lineEdit)

            #self.Layout.insertWidget(self.Layout.count(), self.addButton)
            self.inputs.append(lineEdit)
            self.labels.append(label)

        self.Layout.addWidget(self.addButton)

    def closeEvent(self, event):
        print("clearing stuff")

        for i in reversed(range(self.Layout.count())): 
            self.Layout.itemAt(i).widget().setParent(None)

        event.accept()


class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel data to QTableWidget")
        self.setBaseSize(700, 700)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        self.addData = QPushButton("Добавить Информацию")
        self.add_row = QPushButton("Добавить линию")
        self.remove_row = QPushButton("Удалить линию")
    
        self.filter_list = QComboBox();

        self.DataWindow = DataWindow()

        self.addData.clicked.connect(self.show_data_window)

        self.add_row.clicked.connect(self.addRow)
        self.remove_row.clicked.connect(self.deleteRow)

        self.selectedRow = 0

        self.search = QLineEdit()
        self.search.textChanged.connect(self.findName)

        self.list = QListWidget()
        self.list.setMaximumSize(300, 70)
        self.list.itemClicked.connect(self.listClick)

        layout.addWidget(self.filter_list)
        layout.addWidget(self.search)
        layout.addWidget(self.addData)
        #layout.addWidget(self.add_row)
        layout.addWidget(self.remove_row)
        layout.addWidget(self.list)
        layout.setAlignment(self.list, Qt.AlignmentFlag.AlignCenter)
        
        path = "./data.xlsx"
        self.workbook = openpyxl.load_workbook(path)
        self.temp_sheets = {}

        for name in self.workbook.sheetnames:
            tableWidget = QTableWidget()
            #tableWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            #tableWidget.customContextMenuRequested.connect(self.contextMenu)

            tableWidget.cellClicked.connect(self.cellClicked)

            self.temp_sheets[self.workbook[name]] = tableWidget
            tableWidget.hide()

            layout.addWidget(tableWidget)

        self.currentWidget :QTableWidget = self.temp_sheets[next(iter(self.temp_sheets))]

        self.load_data()

    def show_data_window(self):
        self.DataWindow.show()
        self.DataWindow.load()

    def findName(self):
        name = self.search.text().lower()
        found = False

        for row in range(self.currentWidget.rowCount()):
            for column in range(self.currentWidget.columnCount()):
                header = self.currentWidget.horizontalHeaderItem(column)
                if header != None:
                    if header.text() == self.filter_list.currentText(): 
                        item = self.currentWidget.item(row, column)
                        try:
                            found = name in item.text().lower()
                            self.currentWidget.setRowHidden(row, not found)
                            if found:
                                break
                        except AttributeError:
                            pass

    def listClick(self, item :QListWidgetItem):
        for sheet, widget in self.temp_sheets.items():
            widget.hide()

        self.temp_sheets[self.workbook[item.text()]].show()
        self.currentWidget = self.temp_sheets[self.workbook[item.text()]]

        labels = []

        for c in range(self.currentWidget.columnCount()):
            label = self.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
    
        self.filter_list.clear()

        for i in labels:
            self.filter_list.addItem(i)

    def load_data(self):

        for name in self.workbook.sheetnames:
            self.list.addItem(name)
            self.loadSheet(self.workbook[name])

    
    def loadSheet(self, sheet):

        self.temp_sheets[sheet].setRowCount(sheet.max_row)
        self.temp_sheets[sheet].setColumnCount(sheet.max_column)
        
        list_values = list(sheet.values)
        self.temp_sheets[sheet].setHorizontalHeaderLabels(list_values[0])
        self.temp_sheets[sheet].setSortingEnabled(True)

        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                item = TableWidgetItem(str(value))
                #item.setData(Qt.ItemDataRole.EditRole, str(value))
                self.temp_sheets[sheet].setItem(row_index , col_index, item)
               
                col_index += 1
            row_index += 1

    def addRow(self):
        self.currentWidget.insertRow(self.selectedRow)

        return self.selectedRow

    def deleteRow(self, *args):
        print(self.selectedRow)
        self.currentWidget.removeRow(self.selectedRow)
        print(*args)

    def cellClicked(self, row, column):
        self.selectedRow = row
        


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Window()
    widget = Main()

    window.setCentralWidget(widget)
    widget.show()
    window.show()
    
    app.exec()